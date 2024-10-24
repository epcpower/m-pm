from __future__ import (
    annotations,
)  # See PEP 563, check to remove in future Python version higher than 3.7
import re
import attr
import decimal
import openpyxl
import pathlib
import typing
import glob
import uuid
import mpm.mpm_helper
import epyqlib.treenode
import epyqlib.utils.general
from natsort import natsorted
from tqdm import tqdm

EXCEL_COLUMN_LETTERS = [c for c in "ABCDEFGHIJKLMNOPQRSTU"]
PMVS_UUID_TO_DECIMAL_LIST = typing.List[typing.Dict[uuid.UUID, decimal.Decimal]]

# The parameter query prefix on a large portion of the CAN parameter paths.
PARAMETER_QUERY_PREFIX = "ParameterQuery -> "
# The parameters prefix on a large portion of the parameter paths.
PARAMETERS_PREFIX = "Parameters -> "
TABLES_TREE_STR = "Tables -> Tree"
FILTER_GROUPS = [
    "2. DC",
    "9. Simulation Mode",
    "A. ABB",
    "B. Other -> Authorization",
    "B. Other -> Debug",
]
CELL_SIDE = openpyxl.styles.Side(border_style="thin", color="000000")
CELL_BORDER = openpyxl.styles.Border(
    top=CELL_SIDE, left=CELL_SIDE, right=CELL_SIDE, bottom=CELL_SIDE
)
CELL_FONT = openpyxl.styles.Font(size=8)
# All values are stored as text to have consistent left alignment
NUMBER_FORMAT_TEXT = openpyxl.styles.numbers.FORMAT_TEXT
NUMBERED_VARIANT_PATTERN = r"_(0[2-9]|1[0-9]|20)$"
MIN_MANUAL_COLUMN_COUNT = 5  # Name, access level, min, max, default

builders = epyqlib.utils.general.TypeMap()


@attr.s
class Fields(mpm.mpm_helper.FieldsInterface):
    """The fields defined for a given row in the output XLS file."""

    parameter_name = attr.ib(default=None, type=typing.Union[str, bool])
    can_parameter_name = attr.ib(default=None, type=typing.Union[str, bool])
    description = attr.ib(default=None, type=typing.Union[str, bool])
    access_level = attr.ib(default=None, type=typing.Union[str, bool, int])
    units = attr.ib(default=None, type=typing.Union[str, bool])
    can_path = attr.ib(default=None, type=typing.Union[str, bool])
    parameter_path = attr.ib(default=None, type=typing.Union[str, bool])
    enumerator_list = attr.ib(default=None, type=typing.Union[str, bool])
    parameter_uuid = attr.ib(default=None, type=typing.Union[str, bool])
    epyq_can_parameter_name = attr.ib(default=None, type=typing.Union[str, bool])
    minimum = attr.ib(default=None, type=typing.Union[str, bool, decimal.Decimal])
    maximum = attr.ib(default=None, type=typing.Union[str, bool, decimal.Decimal])
    defaults = attr.ib(
        default=[], type=typing.List[typing.Union[str, bool, decimal.Decimal]]
    )


field_names = Fields(
    parameter_name="Parameter Name",
    can_parameter_name="CAN Parameter Name",
    description="Description",
    units="Units",
    access_level="Access Level",
    can_path="CAN Path",
    parameter_path="Parameter Path",
    enumerator_list="Enumerator List",
    parameter_uuid="Parameter UUID",
    epyq_can_parameter_name="EPyQ CAN Parameter Name",
    minimum="Minimum",
    maximum="Maximum",
    defaults=[],
)


def create_pmvs_uuid_to_value_list(
    pmvs_path: pathlib.Path,
) -> PMVS_UUID_TO_DECIMAL_LIST:
    """
    Creates the pmvs_uuid_to_value_list,
    which is a list of dict's for each pmvs,
    each containing a key -> value of UUID to decimal value

    Args:
        pmvs_path: directory path to the pmvs files

    Returns:
        list of PMVS UUID to decimal dict's
    """
    pmvs_files = glob.glob(f"{pmvs_path}/*.pmvs")
    pmvs_list = []
    for pmvs_file in pmvs_files:
        pmvs = epyqlib.pm.valuesetmodel.loadp(pmvs_file)
        pmvs_list.append(pmvs)
        field_names.defaults.append(pmvs.path.stem)

    pmvs_uuid_to_value_list = []
    for pmvs in pmvs_list:
        pmvs_uuid_to_value = {}
        for child in pmvs.model.root.children:
            pmvs_uuid_to_value.update({child.parameter_uuid: child.value})
        pmvs_uuid_to_value_list.append(pmvs_uuid_to_value)

    return pmvs_uuid_to_value_list


def export(
    path: pathlib.Path,
    can_model: epyqlib.attrsmodel.Model,
    pmvs_path: pathlib.Path,
    column_filter: mpm.mpm_helper.FieldsInterface = None,
) -> None:
    """
    Generate the CAN model parameter data in Excel format (.xlsx).

    Args:
        path: path and filename for .xlsx file
        can_model: CAN model
        parameters_model: parameters model
        pmvs_path: directory path to the pmvs files
        column_filter: columns to be output to .xls file

    Returns:

    """
    pmvs_uuid_to_value_list = create_pmvs_uuid_to_value_list(pmvs_path)

    if column_filter is None:
        column_filter = mpm.mpm_helper.attr_fill(Fields, True)

    builder = mpm.cantoxlsx.builders.wrap(
        wrapped=can_model.root,
        parameter_uuid_finder=can_model.node_from_uuid,
        column_filter=column_filter,
        pmvs_uuid_to_value_list=pmvs_uuid_to_value_list,
    )

    workbook = builder.gen()

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


@builders(mpm.canmodel.Root)
@attr.s
class Root:
    """Excel spreadsheet generator for the CAN Root class."""

    wrapped = attr.ib(type=mpm.canmodel.Root)
    column_filter = attr.ib(type=Fields)
    parameter_uuid_finder = attr.ib(default=None, type=typing.Callable)
    pmvs_uuid_to_value_list = attr.ib(default=None, type=PMVS_UUID_TO_DECIMAL_LIST)

    def gen(self) -> openpyxl.workbook.workbook.Workbook:
        """
        Excel spreadsheet generator for the CAN Root class.

        Returns:
            workbook: generated Excel workbook
        """
        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)
        worksheet = workbook.create_sheet("Parameters")
        worksheet.append(field_names.as_expanded_list(self.column_filter))

        unsorted_rows = []
        for child in self.wrapped.children:
            rows = builders.wrap(
                wrapped=child,
                parameter_uuid_finder=self.parameter_uuid_finder,
                pmvs_uuid_to_value_list=self.pmvs_uuid_to_value_list,
            ).gen()

            for row in rows:
                unsorted_rows.append(row)

        sorted_rows = natsorted(unsorted_rows, key=lambda x: x.parameter_path)

        for row in sorted_rows:
            worksheet.append(row.as_expanded_list(self.column_filter))

        return workbook


@builders(mpm.canmodel.Signal)
@attr.s
class Signal:
    """Excel spreadsheet generator for the CAN Signal class."""

    wrapped = attr.ib(type=mpm.canmodel.Signal)
    parameter_uuid_finder = attr.ib(default=None, type=typing.Callable)
    pmvs_uuid_to_value_list = attr.ib(default=None, type=PMVS_UUID_TO_DECIMAL_LIST)

    def gen(self) -> typing.List[Fields]:
        """
        Excel spreadsheet generator for the CAN Signal class.

        Returns:
            list of a single Fields row for Signal
        """
        if self.wrapped.parameter_uuid:
            row = Fields(defaults=[])
            parameter = self.parameter_uuid_finder(self.wrapped.parameter_uuid)
            self._set_pmvs_row_defaults(
                row, self.pmvs_uuid_to_value_list, self.wrapped.parameter_uuid
            )
            row.can_parameter_name = self.wrapped.name
            row.parameter_name = parameter.name
            row.parameter_uuid = str(self.wrapped.parameter_uuid)
            if isinstance(
                parameter,
                (
                    epyqlib.pm.parametermodel.Parameter,
                    epyqlib.pm.parametermodel.TableArrayElement,
                ),
            ):
                row.units = parameter.units
                row.description = parameter.comment
                access_level = self.parameter_uuid_finder(parameter.access_level_uuid)
                row.access_level = access_level.name
                if parameter.minimum is not None:
                    row.minimum = parameter.minimum
                if parameter.maximum is not None:
                    row.maximum = parameter.maximum

                if self.wrapped.enumeration_uuid is not None:
                    enumeration = self.parameter_uuid_finder(
                        self.wrapped.enumeration_uuid
                    )
                    child_enum_list = []
                    for child_enum in enumeration.children:
                        child_enum_list.append(
                            f"[{child_enum.value}] {child_enum.name}; "
                        )
                    row.enumerator_list = " ".join(child_enum_list)

            parameter_path_list = self._generate_path_list(parameter)
            row.parameter_path = " -> ".join(parameter_path_list)
            can_path_list = self._generate_path_list(self.wrapped)
            can_path_joined = " -> ".join(can_path_list)
            row.can_path = can_path_joined
            can_parameter_path = can_path_joined
            if can_parameter_path.startswith(PARAMETER_QUERY_PREFIX):
                # Chop off the parameter query prefix to match what is seen in the EPyQ parameters tab.
                can_parameter_path = can_parameter_path[len(PARAMETER_QUERY_PREFIX) :]
                # Replace '->' with nothing to match EPyQ name. This is for table parameters.
                can_parameter_path = can_parameter_path.replace(" -> ", "")
            row.epyq_can_parameter_name = f"{can_parameter_path}:{self.wrapped.name}"

            return [row]
        return []

    @staticmethod
    def _generate_path_list(node: epyqlib.treenode.TreeNode) -> typing.List[str]:
        """
        Generate the node's path list.

        Args:
            node: tree node (from CAN model or Parameters model)

        Returns:
            node's path list
        """
        path_list = []
        node_parent = node
        while True:
            if node_parent.tree_parent is not None:
                path_list.insert(0, node_parent.tree_parent.name)
                node_parent = node_parent.tree_parent
            else:
                break
        if len(path_list) > 1:
            # Remove the unnecessary Parameters root element.
            path_list.pop(0)

        return path_list

    @staticmethod
    def _set_pmvs_row_defaults(
        row: Fields,
        pmvs_uuid_to_value_list: PMVS_UUID_TO_DECIMAL_LIST,
        parameter_uuid: uuid.UUID,
    ) -> None:
        """
        Local method to set the PMVS defaults for the given Fields row.

        Args:
            row: single Fields row for Signal
            pmvs_uuid_to_value_list: list of PMVS UUID to decimal dict's
            parameter_uuid: parameter UUID

        Returns:

        """
        for pmvs in pmvs_uuid_to_value_list:
            value = pmvs.get(parameter_uuid)
            row.defaults.append(value)


@builders(mpm.canmodel.CanTable)
@builders(mpm.canmodel.Message)
@builders(mpm.canmodel.Multiplexer)
@builders(mpm.canmodel.MultiplexedMessage)
@builders(mpm.canmodel.MultiplexedMessageClone)
@attr.s
class GenericNode:
    """Excel spreadsheet generator for various CAN model classes."""

    wrapped = attr.ib(
        type=typing.Union[
            mpm.canmodel.CanTable,
            mpm.canmodel.Message,
            mpm.canmodel.Multiplexer,
            mpm.canmodel.MultiplexedMessage,
            mpm.canmodel.MultiplexedMessageClone,
        ]
    )
    parameter_uuid_finder = attr.ib(default=None, type=typing.Callable)
    pmvs_uuid_to_value_list = attr.ib(default=None, type=PMVS_UUID_TO_DECIMAL_LIST)

    def gen(self) -> typing.List[Fields]:
        """
        Excel spreadsheet generator for various CAN model classes.

        Returns:
            list of Fields rows
        """
        output_list = []
        for child in self.wrapped.children:
            frame = builders.wrap(
                wrapped=child,
                parameter_uuid_finder=self.parameter_uuid_finder,
                pmvs_uuid_to_value_list=self.pmvs_uuid_to_value_list,
            ).gen()
            output_list.extend(frame)

        return output_list


def format_for_manual(
    input_path: pathlib.Path,
) -> None:
    """
    Translate the CAN model parameter data to formatted Excel format (.xlsx)
    for the purpose of insertion into the controls manual documentation.

    Args:
        input_path: path and filename for input .xlsx file

    Returns:

    """
    input_workbook = openpyxl.load_workbook(filename=input_path)
    input_worksheet = input_workbook.active
    input_worksheet_col_count = input_worksheet.max_column

    output_path = input_path.with_name(
        input_path.stem + "_for_manual" + input_path.suffix
    )
    print(f"input spreadsheet: {input_path}")
    print(f"output spreadsheet: {output_path}")
    print("Be patient. Generation of the output spreadsheet takes a long time.")

    output_workbook = openpyxl.Workbook()
    output_workbook.remove(output_workbook.active)
    output_worksheet = output_workbook.create_sheet("Parameters")

    # Perform all filtering activities.
    filtered_rows = []
    for row in input_worksheet.iter_rows(min_row=2, max_col=input_worksheet_col_count):
        # Only output parameters that are in EPyQ.
        parameter_path = row[6].value
        if not parameter_path.startswith(PARAMETERS_PREFIX):
            continue

        # Filter out parameter groups in FILTER_GROUPS list.
        filter_out = False
        for group_parameter_filter in FILTER_GROUPS:
            if parameter_path.startswith(PARAMETERS_PREFIX + group_parameter_filter):
                filter_out = True
        if filter_out:
            continue

        # Only allow access levels: Service_Tech and Service_Eng.
        access_level_out = row[3].value
        if access_level_out in ["Service_Tech", "Service_Eng"]:
            filtered_rows.append(row)

    # Track the current row in the output worksheet.
    current_row = 1

    # Display a progress bar since the generation of the output takes a long time.
    with tqdm(total=len(filtered_rows)) as progress_bar:
        current_parameter_path = ""
        entered_tables_section = False

        for row in filtered_rows:
            is_numbered_variant = False

            parameter_path = row[6].value
            if not parameter_path.startswith(PARAMETERS_PREFIX):
                # Only output parameters that are in EPyQ.
                continue
            description_out = row[2].value
            access_level_out = row[3].value
            units_out = row[4].value
            # can_path = row[5].value
            # parameter_path = row[6].value
            enumerator_list = row[7].value
            # uuid = row[8].value
            parameter_name_out = row[9].value
            minimum_out = row[10].value
            maximum_out = row[11].value
            defaults_out = []
            for col in row[12:]:
                if col.value != None:
                    defaults_out.append(f"{col.value}")
                else:
                    defaults_out.append("")

            # is_numbered_variant is necessary to distinguish parameters that are similarly named
            # (differ by numbers) from those that aren't (differ by word(s)) since both have
            # entered_table_section and all_defaults_same set to True
            is_numbered_variant = (
                True
                if re.search(NUMBERED_VARIANT_PATTERN, parameter_name_out)
                else False
            )

            if units_out:
                # If there is a units value, append it to the end of the numeric default value.
                if minimum_out is not None:
                    minimum_out = f"{minimum_out} {units_out}"
                if maximum_out is not None:
                    maximum_out = f"{maximum_out} {units_out}"
                for i in range(len(defaults_out)):
                    if defaults_out[i] != "":
                        defaults_out[i] = f"{defaults_out[i]} {units_out}"

            # Discover if all the product defaults are equal.
            all_defaults_same = len(set(defaults_out)) == 1

            # +1 for parameter name
            column_count = max(MIN_MANUAL_COLUMN_COUNT, len(defaults_out) + 1)

            # If applicable, add a header description for a set of parameters.
            # Chop off the parameters prefix to match the path that is seen in the EPyQ parameters tab.
            parameter_path_to_check = parameter_path[len(PARAMETERS_PREFIX) :]
            if parameter_path_to_check != current_parameter_path:
                # Add the parameter path for this section of parameters.
                current_parameter_path = parameter_path_to_check
                output_worksheet.append([current_parameter_path])

                # Merge cells for header description.
                output_worksheet.merge_cells(
                    start_row=current_row,
                    start_column=1,
                    end_row=current_row,
                    end_column=column_count,
                )

                # Set the font size for header description.
                for col in EXCEL_COLUMN_LETTERS[:column_count]:
                    output_worksheet[col + str(current_row)].font = CELL_FONT

                current_row += 1
                # Reset the tables section logic.
                entered_tables_section = False

            # Track the rows used for each parameter entry.
            rows_used = 0

            if entered_tables_section and is_numbered_variant:
                # Different table defaults for different products would need to keep track of all
                # table parameters and add the header row with product variants if the defaults differ
                # in any of them. So far the default is the same for all products so this is not
                # yet implemented.
                assert (
                    all_defaults_same
                ), "Different defaults for table parameters has not been implemented"
                # Output single Default cells section for additional table row.
                row = (
                    [parameter_name_out, access_level_out]
                    + (column_count - MIN_MANUAL_COLUMN_COUNT) * [""]
                    + [
                        minimum_out,
                        maximum_out,
                        defaults_out[0],
                    ]
                )
                output_worksheet.append(row)
                rows_used += 1

                # Set horizontal & vertical alignment for parameter name.
                output_worksheet[
                    "A" + str(current_row)
                ].alignment = openpyxl.styles.alignment.Alignment(
                    horizontal="left", vertical="top"
                )

                # Merge access level; minimum, maximum and default stay as 1 column
                output_worksheet.merge_cells(
                    start_row=current_row,
                    start_column=2,
                    end_row=current_row,
                    end_column=column_count - 3,
                )

            else:
                # Check and set if this parameter is the start of table rows section.
                entered_tables_section = TABLES_TREE_STR in parameter_path

                # Add the parameter name and description cells.
                output_worksheet.append([parameter_name_out, description_out])
                rows_used += 1

                if enumerator_list:
                    # Add the enumerator list cell / row.
                    output_worksheet.append(["", enumerator_list])
                    rows_used += 1

                if all_defaults_same:
                    row1 = (
                        ["", field_names.access_level]
                        + (column_count - MIN_MANUAL_COLUMN_COUNT) * [""]
                        + [field_names.minimum, field_names.maximum, "Default"]
                    )
                    row2 = (
                        ["", access_level_out]
                        + (column_count - MIN_MANUAL_COLUMN_COUNT) * [""]
                        + [minimum_out, maximum_out, defaults_out[0]]
                    )
                    # Output single Default cells section.
                    output_worksheet.append(row1)
                    output_worksheet.append(row2)
                    rows_used += 2
                else:
                    # Output multiple Default cells sections, +1 for no default column
                    row1 = (
                        ["", field_names.access_level]
                        + (column_count - MIN_MANUAL_COLUMN_COUNT + 1) * [""]
                        + [field_names.minimum, field_names.maximum]
                    )
                    row2 = (
                        ["", access_level_out]
                        + (column_count - MIN_MANUAL_COLUMN_COUNT + 1) * [""]
                        + [minimum_out, maximum_out]
                    )
                    output_worksheet.append(row1)
                    output_worksheet.append(row2)
                    output_worksheet.append([""] + field_names.defaults)
                    output_worksheet.append([""] + defaults_out)
                    rows_used += 4

                # Merge cells for parameter name.
                output_worksheet.merge_cells(
                    start_row=current_row,
                    start_column=1,
                    end_row=current_row + rows_used - 1,
                    end_column=1,
                )

                # Set horizontal & vertical alignment for parameter name.
                output_worksheet[
                    "A" + str(current_row)
                ].alignment = openpyxl.styles.alignment.Alignment(
                    horizontal="left", vertical="top"
                )

                # Set alignment of description to wrap text.
                for col in EXCEL_COLUMN_LETTERS[1:column_count]:
                    output_worksheet[
                        col + str(current_row)
                    ].alignment = openpyxl.styles.alignment.Alignment(wrap_text=True)

                # Merge cells for description.
                output_worksheet.merge_cells(
                    start_row=current_row,
                    start_column=2,
                    end_row=current_row,
                    end_column=column_count,
                )

                if enumerator_list:
                    # Merge cells for enumerator list cell.
                    output_worksheet.merge_cells(
                        start_row=current_row + 1,
                        start_column=2,
                        end_row=current_row + 1,
                        end_column=column_count,
                    )

                if all_defaults_same:
                    # Merge access level; minimum, maximum and default stay as 1 column
                    if enumerator_list:
                        output_worksheet.merge_cells(
                            start_row=current_row + 2,
                            start_column=2,
                            end_row=current_row + 2,
                            end_column=column_count - 3,
                        )
                        output_worksheet.merge_cells(
                            start_row=current_row + 3,
                            start_column=2,
                            end_row=current_row + 3,
                            end_column=column_count - 3,
                        )
                    else:
                        output_worksheet.merge_cells(
                            start_row=current_row + 1,
                            start_column=2,
                            end_row=current_row + 1,
                            end_column=column_count - 3,
                        )
                        output_worksheet.merge_cells(
                            start_row=current_row + 2,
                            start_column=2,
                            end_row=current_row + 2,
                            end_column=column_count - 3,
                        )
                else:
                    # Merge access level; minimum and maximum stay as 1 column; no default column.
                    # The product specific defaults each get their own column.
                    if enumerator_list:
                        output_worksheet.merge_cells(
                            start_row=current_row + 2,
                            start_column=2,
                            end_row=current_row + 2,
                            end_column=column_count - 2,
                        )
                        output_worksheet.merge_cells(
                            start_row=current_row + 3,
                            start_column=2,
                            end_row=current_row + 3,
                            end_column=column_count - 2,
                        )
                    else:
                        output_worksheet.merge_cells(
                            start_row=current_row + 1,
                            start_column=2,
                            end_row=current_row + 1,
                            end_column=column_count - 2,
                        )
                        output_worksheet.merge_cells(
                            start_row=current_row + 2,
                            start_column=2,
                            end_row=current_row + 2,
                            end_column=column_count - 2,
                        )

            # Set the font size and border for non header description rows.
            for style_row in range(current_row, current_row + rows_used):
                for col in EXCEL_COLUMN_LETTERS[:column_count]:
                    output_worksheet[col + str(style_row)].font = CELL_FONT
                    output_worksheet[col + str(style_row)].border = CELL_BORDER
                    output_worksheet[
                        col + str(style_row)
                    ].number_format = NUMBER_FORMAT_TEXT

            # Update the current row with the number of rows used plus one to go to the next row.
            current_row += rows_used
            progress_bar.update(1)

    output_workbook.save(output_path)
